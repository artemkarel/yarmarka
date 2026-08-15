# -*- coding: utf-8 -*-
"""Таблица покрытия: населённые пункты × категории событий (есть в базе / нет).

Строки: все муниципалитеты Нижегородской и Свердловской областей (полные
официальные перечни) + все города из базы. Столбцы: ярмарки коммерческие,
сельхозярмарки, праздники/дни городов/фестивали. Выход: xlsx.
Запуск: python scripts/coverage_report.py [выходной_файл.xlsx]
"""
import os
import re
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402

from app import db  # noqa: E402

OUT = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser(
    "~/Desktop/покрытие_мероприятий.xlsx")

# Официальные перечни муниципальных образований (города/округа/районы)
NIZHNY = [
    "Нижний Новгород", "Арзамас", "Бор", "Дзержинск", "Кстово", "Выкса", "Саров",
    "Павлово", "Балахна", "Заволжье", "Городец", "Кулебаки", "Лысково", "Семёнов",
    "Богородск", "Навашино", "Лукоянов", "Сергач", "Чкаловск", "Шахунья", "Ветлуга",
    "Урень", "Первомайск", "Перевоз", "Княгинино", "Володарск", "Ворсма", "Горбатов",
    "Ардатовский р-н", "Большеболдинский р-н", "Большемурашкинский р-н",
    "Бутурлинский р-н", "Вадский р-н", "Варнавинский р-н", "Вачский р-н",
    "Вознесенский р-н", "Воротынский р-н", "Воскресенский р-н", "Гагинский р-н",
    "Дальнеконстантиновский р-н", "Дивеевский р-н", "Ковернинский р-н",
    "Краснобаковский р-н", "Краснооктябрьский р-н", "Пильнинский р-н",
    "Починковский р-н", "Сеченовский р-н", "Сокольский р-н", "Сосновский р-н",
    "Спасский р-н", "Тонкинский р-н", "Тоншаевский р-н", "Шарангский р-н",
    "Шатковский р-н",
]
SVERDLOVSK = [
    "Екатеринбург", "Нижний Тагил", "Каменск-Уральский", "Первоуральск", "Серов",
    "Новоуральск", "Верхняя Пышма", "Асбест", "Полевской", "Ревда", "Краснотурьинск",
    "Берёзовский", "Лесной", "Верхняя Салда", "Качканар", "Красноуфимск", "Алапаевск",
    "Ирбит", "Реж", "Тавда", "Сухой Лог", "Артёмовский", "Богданович", "Карпинск",
    "Заречный", "Камышлов", "Красноуральск", "Кировград", "Невьянск", "Североуральск",
    "Среднеуральск", "Нижняя Тура", "Сысерть", "Кушва", "Талица", "Арамиль",
    "Дегтярск", "Туринск", "Волчанск", "Ивдель", "Нижние Серги", "Михайловск",
    "Нижняя Салда", "Верхний Тагил", "Верхняя Тура", "Верхотурье", "Новая Ляля",
    "Байкаловский р-н", "Ачитский округ", "Артинский округ", "Белоярский округ",
    "Гаринский округ", "Горноуральский округ", "Каменский округ", "Красноуфимский округ",
    "Пышминский округ", "Слободо-Туринский р-н", "Сосьвинский округ", "Таборинский р-н",
    "Тугулымский округ", "Туринская Слобода", "Шалинский округ", "Верхнее Дуброво",
    "Малышева", "Рефтинский", "Пелым", "Бисерть", "Верх-Нейвинский", "Свободный",
    "Уральский", "Староуткинск",
]


def norm(s):
    """Грубая нормализация для сопоставления города из базы со справочником."""
    s = str(s or "").lower().replace("ё", "е")
    s = re.sub(r"[«»\"().,/]", " ", s)
    s = re.sub(r"\b(муниципальн\w+|образование|округ|городской|район|р-н|город|г|пос|"
               r"пгт|село|с|деревня|д|мо|го|обл|область|и других?|люб\w+)\b", " ", s)
    # «новоуральский» → «новоуральск»: прилагательные округов к основе города
    s = re.sub(r"ск(ий|ое|ая|ого|ому|им|ом)\b", "ск", s)
    return re.sub(r"\s+", " ", s).strip()


def main():
    conn = db.get()
    # категория -> множество нормализованных городов из базы
    cats = {
        "commerce": "etype='Ярмарка коммерческая'",
        "agro": "etype='Сельхозярмарка'",
        "other": "etype IN ('Праздник','День города/села','Фестиваль','Другое')",
    }
    db_cities = {}
    for key, cond in cats.items():
        vals = set()
        for r in conn.execute(f"SELECT DISTINCT city FROM events WHERE {cond}"):
            vals.add(norm(r["city"]))
        db_cities[key] = vals
    # сельхоз-точки тоже считаются сельхоз-покрытием
    for r in conn.execute("SELECT DISTINCT city FROM points WHERE ptype='Сельхозярмарка'"):
        db_cities["agro"].add(norm(r["city"]))

    counts = {}
    for r in conn.execute("SELECT city, COUNT(*) n FROM events GROUP BY city"):
        k = norm(r["city"])
        counts[k] = counts.get(k, 0) + r["n"]

    def eqw(x, y):
        """Слова равны или одно — основа другого («асбест» ~ «асбестовск»)."""
        return x == y or (len(x) >= 5 and len(y) >= 5 and
                          (x.startswith(y) or y.startswith(x)))

    def same(a, b):
        if not a or not b:
            return False
        aw, bw = a.split(), b.split()
        small, big = (aw, bw) if len(aw) <= len(bw) else (bw, aw)
        return all(any(eqw(w, v) for v in big) for w in small)

    def hit(cat, name):
        n = norm(name)
        return any(same(n, c) for c in db_cities[cat])

    rows = []
    seen = set()
    for region, places in (("Нижегородская обл.", NIZHNY), ("Свердловская обл.", SVERDLOVSK)):
        for p in places:
            rows.append((region, p))
            seen.add(norm(p))
    # города из базы, которых нет в справочнике (другие регионы: exponet и т.п.)
    extra = set()
    for r in conn.execute("SELECT DISTINCT city FROM events WHERE city != ''"):
        n = norm(r["city"])
        if n and not any(same(n, s) for s in seen):
            extra.add(r["city"])
    for city in sorted(extra):
        rows.append(("Другие регионы", city))

    wb = Workbook()
    ws = wb.active
    ws.title = "Покрытие"
    head = ["Регион", "Населённый пункт", "Ярмарки коммерческие", "Сельхозярмарки",
            "Праздники и дни городов", "Всего событий в базе"]
    ws.append(head)
    bold = Font(bold=True)
    for c in ws[1]:
        c.font = bold
    green = PatternFill("solid", start_color="D8F0DC")
    red = PatternFill("solid", start_color="F8DCDC")
    center = Alignment(horizontal="center")
    for region, place in rows:
        n = norm(place)
        marks = [hit("commerce", place), hit("agro", place), hit("other", place)]
        total = sum(v for k, v in counts.items() if same(k, n))
        ws.append([region, place] + ["ЕСТЬ" if m else "нет" for m in marks] + [total or ""])
        for j, m in enumerate(marks):
            cell = ws.cell(row=ws.max_row, column=3 + j)
            cell.fill = green if m else red
            cell.alignment = center
        ws.cell(row=ws.max_row, column=6).alignment = center
    widths = [20, 30, 21, 16, 22, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    wb.save(OUT)
    print("строк:", len(rows), "->", OUT)


if __name__ == "__main__":
    main()
