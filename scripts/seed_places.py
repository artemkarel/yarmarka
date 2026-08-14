# -*- coding: utf-8 -*-
"""Импорт мероприятий и точек из файла «yarmarki_rossiya.xlsx».

Запуск: python scripts/seed_places.py [путь_к_xlsx]
Строки с конкретной датой попадают в календарь мероприятий; постоянные и
еженедельные площадки без даты — в справочник точек. Повторный запуск
пропускает уже загруженные записи.
"""
import os
import re
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl  # noqa: E402

from app import db, services  # noqa: E402

XLSX = sys.argv[1] if len(sys.argv) > 1 else \
    os.path.expanduser("~/Downloads/yarmarki_rossiya.xlsx")

EVENT_SHEETS = ["Праздники и Дни городов", "Сельхозярмарки и базары", "Коммерческие ярмарки"]
POINT_SHEETS = ["Рынки", "Магазины и ТЦ"]

CITY_PREFIXES = ("г.о.", "р.п.", "пос.", "п.", "с.", "д.", "г.")

RE_RANGE_FULL = re.compile(
    r"(\d{1,2})\.(\d{1,2})\.(\d{4})\s*[–\-]\s*(\d{1,2})\.(\d{1,2})\.(\d{4})")
RE_RANGE_XMONTH = re.compile(r"(\d{1,2})\.(\d{1,2})\s*[–\-]\s*(\d{1,2})\.(\d{1,2})\.(\d{4})")
RE_RANGE = re.compile(r"(\d{1,2})\s*[–\-]\s*(\d{1,2})\.(\d{1,2})\.(\d{4})")
RE_DATE = re.compile(r"(\d{1,2})\.(\d{1,2})\.(\d{4})")
MONTHS = [("январ", 1), ("феврал", 2), ("март", 3), ("апрел", 4), ("ма", 5), ("июн", 6),
          ("июл", 7), ("август", 8), ("сентябр", 9), ("октябр", 10), ("ноябр", 11),
          ("декабр", 12)]


def short_city(mun):
    if not mun:
        return ""
    part = str(mun).split(",")[-1].strip()
    changed = True
    while changed:
        changed = False
        for pref in CITY_PREFIXES:
            if part.startswith(pref):
                part = part[len(pref):].strip()
                changed = True
    return part or str(mun).strip()


def parse_period(raw):
    """-> (date_from, date_to) или None, если конкретной даты нет."""
    if not raw:
        return None
    s = str(raw).strip()
    m = RE_RANGE_FULL.search(s)
    if m:
        d1, m1, y1, d2, m2, y2 = map(int, m.groups())
        a, b = f"{y1:04d}-{m1:02d}-{d1:02d}", f"{y2:04d}-{m2:02d}-{d2:02d}"
        return (a, b) if b >= a else (a, a)
    m = RE_RANGE_XMONTH.search(s)
    if m:
        d1, m1, d2, m2, y = map(int, m.groups())
        a, b = f"{y:04d}-{m1:02d}-{d1:02d}", f"{y:04d}-{m2:02d}-{d2:02d}"
        if b < a:  # переход через новый год, напр. 28.12–05.01.2027
            a = f"{y - 1:04d}-{m1:02d}-{d1:02d}"
        return (a, b)
    m = RE_RANGE.search(s)
    if m:
        d1, d2, mo, y = map(int, m.groups())
        return (f"{y:04d}-{mo:02d}-{d1:02d}", f"{y:04d}-{mo:02d}-{d2:02d}")
    m = RE_DATE.search(s)
    if m:
        d, mo, y = map(int, m.groups())
        return (f"{y:04d}-{mo:02d}-{d:02d}",) * 2
    low = s.lower()
    ym = re.search(r"(\d{4})", s)
    if ym:
        for pref, num in MONTHS:
            if low.startswith(pref):
                y = int(ym.group(1))
                return (f"{y:04d}-{num:02d}-01", f"{y:04d}-{num:02d}-28")
    return None


def build_comment(*parts):
    out = [str(p).strip() for p in parts if p and str(p).strip() not in ("—", "-", "None")]
    return " • ".join(out)[:500] or None


def main():
    if not os.path.exists(XLSX):
        raise SystemExit(f"Файл не найден: {XLSX}")
    conn = db.get()
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    fake_user = {"id": 0, "role": "admin"}
    ev_added = pt_added = skipped = 0

    def event_exists(name, city, date_from):
        return conn.execute(
            "SELECT 1 FROM events WHERE name=? AND city=? AND date_from=?",
            (name, city, date_from)).fetchone() is not None

    def point_exists(name, city):
        return conn.execute(
            "SELECT 1 FROM points WHERE name=? AND city=?", (name, city)).fetchone() is not None

    def add_point(name, ptype, city, address, comment):
        nonlocal pt_added, skipped
        if point_exists(name, city):
            skipped += 1
            return
        services.point_save(conn, fake_user, None, name, ptype, city, address, None, comment)
        pt_added += 1

    for sheet in EVENT_SHEETS:
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            (region, mun, place, name, channel, etype, fmt, period, freq, cost,
             contact, phone, site, source, status, notes) = (list(row) + [None] * 16)[:16]
            if not name:
                continue
            name = str(name).strip()
            city = short_city(mun)
            dates = parse_period(period)
            comment = build_comment(place if dates else None, contact, phone, cost, site,
                                    None if dates else period)
            if dates:
                if event_exists(name, city, dates[0]):
                    skipped += 1
                    continue
                services.event_save(conn, fake_user, None, name, str(etype or "").strip(),
                                    city, dates[0], dates[1], None, comment)
                ev_added += 1
            else:
                # постоянные/еженедельные площадки без даты — в точки
                add_point(name, str(etype or "Ярмарка").strip(), city,
                          str(place).strip() if place else None, comment)

    for sheet in POINT_SHEETS:
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            (region, mun, place, name, channel, etype, fmt, period, freq, cost,
             contact, phone, site, source, status, notes) = (list(row) + [None] * 16)[:16]
            if not name:
                continue
            name = str(name).strip()
            city = short_city(mun)
            if sheet == "Рынки":
                ptype = "Рынок"
            else:
                ptype = "Магазин" if "магазин" in str(channel or "").lower() else "ТЦ"
            comment = build_comment(fmt, contact, phone, cost, site)
            add_point(name, ptype, city, str(place).strip() if place else None, comment)

    print(f"Мероприятий добавлено: {ev_added}, точек: {pt_added}, пропущено (дубли): {skipped}")


if __name__ == "__main__":
    main()
